import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 读取盘点清单
df = pd.read_excel(r"F:\盘点电脑清单.xlsx", sheet_name="电脑汇总")
print(f"总条数: {len(df)}")

# 清理数据：过滤空行（电脑名和用户名都为空）
mask = df["Computer Name"].notna()
df = df[mask].copy()
df = df.reset_index(drop=True)
print(f"过滤后条数: {len(df)}")

# 映射字段
data = []
for _, row in df.iterrows():
    # 资产名称：优先用 user name，其次 Computer Name
    asset_name = row.get("user name")
    if pd.isna(asset_name) or str(asset_name).strip() == "":
        asset_name = row.get("Computer Name", "")
    if pd.isna(asset_name):
        asset_name = row.get("Computer Name", "")

    # 型号解析品牌
    model = row.get("型号", "")
    brand = ""
    if pd.notna(model):
        model_str = str(model)
        if "HP" in model_str or "HP " in model_str or "ProBook" in model_str or "ProDesk" in model_str or "EliteBook" in model_str:
            brand = "惠普"
        elif "Surface" in model_str:
            brand = "微软"
        elif "ThinkPad" in model_str or "Lenovo" in model_str:
            brand = "联想"
        elif "Dell" in model_str or "戴尔" in model_str:
            brand = "戴尔"
        else:
            brand = "惠普"  # 默认惠普（清单中大部分是HP）
    else:
        brand = "惠普"

    # 序列号：优先用 SN，其次用 资产编码
    sn = row.get("SN", "")
    if pd.isna(sn) or str(sn).strip() == "":
        sn = row.get("资产编码", "")
    if pd.isna(sn) or str(sn).strip() == "":
        sn = row.get("Computer Name", "")  # 最后用电脑名

    # 存放位置
    location = row.get("存放位置", "")
    if pd.isna(location) or str(location).strip() == "":
        location = row.get("部门", "未知")

    # 使用日期 -> 采购日期
    purchase_date = row.get("使用日期", None)

    # 资产编号：用 Computer Name
    asset_code = row.get("Computer Name", "")

    data.append({
        "资产编号(可选)": asset_code if pd.notna(asset_code) else "",
        "资产名称(必填)": asset_name if pd.notna(asset_name) else str(asset_code),
        "分类ID(必填)": 2,  # 计算机
        "分类名称(参考)": "计算机",
        "品牌": brand,
        "型号": model if pd.notna(model) else "",
        "序列号(必填)": str(sn) if pd.notna(sn) else str(asset_code),
        "采购价格": "",
        "采购日期(YYYY-MM-DD)": str(purchase_date)[:10] if pd.notna(purchase_date) else "",
        "供应商": "",
        "存放位置(必填)": str(location) if pd.notna(location) else "未知",
        "资产状态(0-3)": 1,  # 已有使用人，设为已领用
        "备注": f"部门:{row.get('部门', '')}, 电脑名:{row.get('Computer Name', '')}" if pd.notna(row.get('部门')) else "",
    })

print(f"准备写入 {len(data)} 条数据")

# 写入 Excel
output_path = r"E:\固定资产系统\资产导入_电脑盘点数据.xlsx"
wb = load_workbook(r"E:\固定资产系统\资产导入模板.xlsx")
ws = wb["导入模板"]

# 清空旧示例数据（保留表头第1行，清空2-1000行）
for row in range(2, 1001):
    for col in range(1, 14):
        ws.cell(row=row, column=col).value = None

# 写入数据
header_fill = PatternFill("solid", start_color="DCE6F1")
alt_fill = PatternFill("solid", start_color="EEF2F8")
thin = Side(border_style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

headers = [ws.cell(row=1, column=c).value for c in range(1, 14)]

for r_idx, row_data in enumerate(data, 2):
    fill = header_fill if r_idx % 2 == 0 else alt_fill
    values = [
        row_data["资产编号(可选)"],
        row_data["资产名称(必填)"],
        row_data["分类ID(必填)"],
        row_data["分类名称(参考)"],
        row_data["品牌"],
        row_data["型号"],
        row_data["序列号(必填)"],
        row_data["采购价格"],
        row_data["采购日期(YYYY-MM-DD)"],
        row_data["供应商"],
        row_data["存放位置(必填)"],
        row_data["资产状态(0-3)"],
        row_data["备注"],
    ]
    for c_idx, val in enumerate(values, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.fill = fill
        cell.border = border
        cell.alignment = center
        cell.font = Font(name="微软雅黑", size=10)

print(f"已写入 {len(data)} 条到 {output_path}")

# 设置列宽
col_widths = [18, 20, 12, 14, 12, 14, 18, 12, 20, 14, 18, 14, 30]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(output_path)
print("保存成功！")
