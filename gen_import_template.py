from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ===== 1. 导入模板 Sheet =====
ws = wb.active
ws.title = "导入模板"

# 表头定义
headers = [
    "资产编号(可选)",
    "资产名称(必填)",
    "分类ID(必填)",
    "分类名称(参考)",
    "品牌",
    "型号",
    "序列号(必填)",
    "采购价格",
    "采购日期(YYYY-MM-DD)",
    "供应商",
    "存放位置(必填)",
    "资产状态(0-3)",
    "备注"
]

# 样式
header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", start_color="4F81BD")
cat_fill = PatternFill("solid", start_color="DCE6F1")
alt_fill = PatternFill("solid", start_color="EEF2F8")
required_font = Font(name="微软雅黑", color="C00000", size=10)
normal_font = Font(name="微软雅黑", size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(border_style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 写表头
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# 示例数据（2行）
examples = [
    ["SRV-001", "戴尔服务器R750", 1, "服务器", "戴尔", "R750", "DELL-001", 25000.00, "2025-01-15", "戴尔中国", "一号机房A区", 0, "核心业务服务器"],
    ["LT-001", "ThinkPad笔记本电脑", 2, "计算机", "联想", "ThinkPad E14", "SN2025001", 8500.00, "2025-03-20", "联想集团", "201办公室", 1, "研发部使用"],
]

for r, row_data in enumerate(examples, 2):
    fill = cat_fill if r % 2 == 0 else alt_fill
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# 第3行空行（供填写）
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=3, column=c)
    cell.border = border
    cell.fill = alt_fill

# 列宽
col_widths = [18, 20, 12, 14, 12, 14, 18, 12, 20, 14, 18, 14, 20]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 冻结首行
ws.freeze_panes = "A2"

# 数据验证：状态列（L列）
dv_status = DataValidation(type="whole", operator="between", formula1="0", formula2="3", allow_blank=True)
dv_status.error = "资产状态只能填 0~3：0=未领用，1=已领用，2=维修中，3=已报废"
dv_status.errorTitle = "输入错误"
ws.add_data_validation(dv_status)
dv_status.add("L4:L1000")

# 分类ID验证（C列）
dv_cat = DataValidation(type="whole", operator="between", formula1="1", formula2="9", allow_blank=True)
dv_cat.error = "分类ID请填写系统中存在的分类ID（目前：1=服务器，2=计算机，3=打印机）"
dv_cat.errorTitle = "输入错误"
ws.add_data_validation(dv_cat)
dv_cat.add("C4:C1000")

# ===== 2. 填写说明 Sheet =====
ws2 = wb.create_sheet("填写说明")

desc_headers = ["字段名", "是否必填", "说明/可选值"]
for col, h in enumerate(desc_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

explanations = [
    ["资产编号", "可选", "不填则系统自动生成，建议格式如 SRV-001"],
    ["资产名称", "必填", "资产的具体名称，如：戴尔服务器R750"],
    ["分类ID", "必填", "1=服务器，2=计算机，3=打印机（可在系统分类管理中查看）"],
    ["分类名称", "可选", "仅供填写参考，系统以分类ID为准"],
    ["品牌", "可选", "如：戴尔、联想、惠普、思科等"],
    ["型号", "可选", "如：R750、ThinkPad E14、LaserJet Pro等"],
    ["序列号", "必填", "资产唯一序列号，不可重复"],
    ["采购价格", "可选", "单位：元，可填小数，如：25000.00"],
    ["采购日期", "可选", "格式：YYYY-MM-DD，如：2025-01-15"],
    ["供应商", "可选", "如：戴尔中国、联想集团"],
    ["存放位置", "必填", "资产物理存放位置，如：一号机房A区、201办公室"],
    ["资产状态", "可选", "0=未领用（默认），1=已领用，2=维修中，3=已报废"],
    ["备注", "可选", "其他补充说明信息"],
]

req_fill = PatternFill("solid", start_color="FFE6E6")
opt_fill = PatternFill("solid", start_color="E6F5E6")

for r, row_data in enumerate(explanations, 2):
    is_req = "必填" in row_data[1]
    fill = req_fill if is_req else opt_fill
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = Font(name="微软雅黑", bold=is_req, color="C00000" if is_req else "000000", size=10)

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 50

# 使用提示
ws2.merge_cells("E2:E16")
tip_cell = ws2.cell(row=2, column=5, value="💡 使用提示：\n\n1. 必填字段已标红，请勿留空\n2. 资产编号不填则系统自动生成\n3. 批量填写后，通过系统「批量导入」功能上传本文件\n4. 分类ID必须与系统中已有分类一致\n5. 序列号不可重复，请确认后再导入")
tip_cell.alignment = Alignment(vertical="top", wrap_text=True)
tip_cell.font = Font(name="微软雅黑", size=10, color="595959")
tip_cell.fill = PatternFill("solid", start_color="FFF2CC")
ws2.column_dimensions["E"].width = 40

# ===== 3. 分类参考 Sheet =====
ws3 = wb.create_sheet("分类参考")
cat_headers = ["分类ID", "分类名称", "说明"]
for col, h in enumerate(cat_headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

categories = [
    [1, "服务器", "各类服务器设备"],
    [2, "计算机", "台式机、笔记本电脑等"],
    [3, "打印机", "各类打印设备"],
]

for r, row_data in enumerate(categories, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        cell.font = normal_font
        cell.fill = cat_fill if r % 2 == 0 else alt_fill

ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 30

# 保存
output_path = r"E:\固定资产系统\资产导入模板.xlsx"
wb.save(output_path)
print(f"已生成：{output_path}")
